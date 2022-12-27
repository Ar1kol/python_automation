from openpyxl import load_workbook
from openpyxl.chart import (
    BarChart,
    Reference,
)
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import calendar

workbook = load_workbook("pivot_table.xlsx")
sheet = workbook["Report"]
month = calendar.month_name[datetime.now().month]

min_column = workbook.active.min_column
max_column = workbook.active.max_column
min_row = workbook.active.min_row
max_row = workbook.active.max_row

#adding total by column
for col in range(min_column + 1, max_column + 1):
    letter = get_column_letter(col)
    sheet[f"{letter}{max_column +1}"] = f"=SUM({letter}{min_row + 1}:{letter}{max_row})"
    sheet[f"{letter}{max_column +1}"].style = "Currency"

#adding chart
bar_chart = BarChart()
data = Reference(sheet, min_col = min_column + 1, max_col = max_column, min_row = min_row, max_row = max_row)
categories = Reference(sheet, min_col = min_column, max_col = min_column, min_row = min_row + 1, max_row = max_row)

bar_chart.add_data(data, titles_from_data = True)
bar_chart.set_categories(categories)
bar_chart.title = "Sales by Product line"
bar_chart.style = 5

sheet.add_chart(bar_chart, "B12")

#adding title and subtitle
sheet["A1"] = "Sales Report"
sheet["A2"] = month
sheet["A1"].font = Font("Ariel", bold=True, size=20)
sheet["A2"].font = Font("Ariel", bold=True, size=10)

workbook.save(f"report-{month}.xlsx")

    

